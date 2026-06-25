from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from django.views import View
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q

from empresas.models import Empresa, Vehiculo, Conductor
from programa.models import VisitaProgramada
from auditorias.models import Auditoria, AccionMejora, Hallazgo


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "seguimiento/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        hoy = timezone.localdate()

        ctx["total_empresas"] = Empresa.objects.filter(activa=True).count()
        ctx["auditorias_en_proceso"] = Auditoria.objects.filter(estado=Auditoria.Estado.EN_PROCESO).count()
        ctx["auditorias_cerradas"] = Auditoria.objects.filter(estado=Auditoria.Estado.CERRADA).count()

        ctx["proximas_visitas"] = (
            VisitaProgramada.objects.filter(estado=VisitaProgramada.Estado.PROGRAMADA, fecha_programada__gte=hoy)
            .select_related("empresa").order_by("fecha_programada")[:8]
        )

        acciones = AccionMejora.objects.exclude(estado=AccionMejora.Estado.CUMPLIDA)
        ctx["acciones_vencidas"] = acciones.filter(fecha_compromiso__lt=hoy).count()
        ctx["acciones_pendientes"] = acciones.filter(fecha_compromiso__gte=hoy).count()
        ctx["acciones_proximas"] = (
            acciones.select_related("hallazgo__auditoria__empresa").order_by("fecha_compromiso")[:8]
        )

        ctx["vehiculos_soat_vencido"] = sum(1 for v in Vehiculo.objects.filter(activo=True) if v.soat_vencido)
        ctx["vehiculos_rtm_vencida"] = sum(1 for v in Vehiculo.objects.filter(activo=True) if v.rtm_vencida)
        ctx["conductores_con_alertas"] = sum(1 for c in Conductor.objects.filter(activo=True) if c.alertas)

        ctx["no_conformidades_abiertas"] = Hallazgo.objects.filter(
            tipo=Hallazgo.Tipo.NO_CONFORMIDAD
        ).exclude(
            acciones_mejora__estado=AccionMejora.Estado.CUMPLIDA
        ).distinct().count()

        return ctx


class MatrizSeguimientoView(LoginRequiredMixin, View):
    template_name = "seguimiento/matriz.html"

    def _queryset(self, request):
        qs = AccionMejora.objects.select_related(
            "hallazgo", "hallazgo__auditoria", "hallazgo__auditoria__empresa", "hallazgo__componente"
        ).all()
        empresa_id = request.GET.get("empresa")
        estado = request.GET.get("estado")
        if empresa_id:
            qs = qs.filter(hallazgo__auditoria__empresa_id=empresa_id)
        if estado:
            qs = qs.filter(estado=estado)
        return qs.order_by("fecha_compromiso")

    def get(self, request):
        from django.shortcuts import render
        acciones = self._queryset(request)
        if request.GET.get("formato") == "xlsx":
            return self._exportar_excel(acciones)
        return render(request, self.template_name, {
            "acciones": acciones,
            "empresas": Empresa.objects.filter(activa=True).order_by("nombre"),
            "estados": AccionMejora.Estado.choices,
        })

    def _exportar_excel(self, acciones):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matriz de seguimiento"
        encabezados = [
            "Empresa", "NIT", "Fecha auditoría", "Tipo hallazgo", "Componente",
            "Hallazgo", "Acción de mejora", "Responsable", "Fecha compromiso",
            "Estado", "Fecha de cierre",
        ]
        ws.append(encabezados)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")

        for acc in acciones:
            h = acc.hallazgo
            a = h.auditoria
            ws.append([
                a.empresa.nombre, a.empresa.nit, a.fecha_inicio.strftime("%Y-%m-%d"),
                h.get_tipo_display(), h.componente.nombre if h.componente else "",
                h.descripcion, acc.accion, acc.responsable,
                acc.fecha_compromiso.strftime("%Y-%m-%d"), acc.get_estado_display(),
                acc.fecha_cierre.strftime("%Y-%m-%d") if acc.fecha_cierre else "",
            ])

        for i, ancho in enumerate([28, 14, 14, 16, 26, 40, 40, 20, 16, 14, 14], start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Matriz_Seguimiento_PESV.xlsx"
        wb.save(response)
        return response
